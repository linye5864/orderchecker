"""
文件处理服务
Excel 文件解析和验证
"""

import os
import uuid
from pathlib import Path
from typing import Optional, List, Tuple, Dict, Any, Literal
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from app.core.config import settings
from app.models.orm_file import UploadedFile
from sqlalchemy.orm import Session


# 文件类型
FileTypeLiteral = Literal["delivery", "platform", "flow"]


class FileProcessingService:
    """文件处理服务"""
    
    def __init__(self, db: Session):
        self.db = db
        self.upload_dir = Path(settings.UPLOAD_DIR)
        self.upload_dir.mkdir(parents=True, exist_ok=True)
    
    async def save_uploaded_file(
        self,
        file_content: bytes,
        original_filename: str,
        file_type: str,
        user_id: Optional[str] = None,
    ) -> UploadedFile:
        """保存上传的文件"""
        # 生成唯一文件名
        file_id = uuid.uuid4()
        extension = Path(original_filename).suffix
        stored_filename = f"{file_id}{extension}"
        file_path = self.upload_dir / stored_filename
        
        # 保存文件
        with open(file_path, "wb") as f:
            f.write(file_content)
        
        # 获取文件大小
        file_size = len(file_content)
        
        # 创建数据库记录
        uploaded_file = UploadedFile(
            id=file_id,
            original_filename=original_filename,
            stored_filename=stored_filename,
            file_path=str(file_path),
            file_size=file_size,
            file_type=file_type,
            mime_type=self._get_mime_type(original_filename),
            uploaded_by=uuid.UUID(user_id) if user_id else None,
        )
        
        self.db.add(uploaded_file)
        self.db.commit()
        self.db.refresh(uploaded_file)
        
        return uploaded_file
    
    def get_file_by_id(self, file_id: uuid.UUID) -> Optional[UploadedFile]:
        """根据 ID 获取文件记录"""
        return self.db.query(UploadedFile).filter(UploadedFile.id == file_id).first()
    
    def detect_file_type(self, file_path: str) -> Optional[str]:
        """检测文件类型"""
        extension = Path(file_path).suffix.lower()
        
        if extension in [".xlsx", ".xls"]:
            # 尝试通过内容检测
            try:
                df = pd.read_excel(file_path, nrows=5)
                columns = [str(c).lower() for c in df.columns]
                
                # 检测配送单
                delivery_markers = ["配送单订单号", "发单运力", "配送状态", "free"]
                if any(marker.lower() in columns for marker in delivery_markers):
                    return "delivery"
                
                # 检测平台账单 (支持多种平台: 达达, 顺丰, 闪送, etc.)
                platform_markers = [
                    "三方订单编号", "第三方订单ID", "第三方订单号",
                    "订单编号", "订单号", "达达订单ID",
                    "订单状态", "状态",
                    "实付金额", "应付金额", "金额",
                    "取消扣款金额", "扣款金额"
                ]
                if any(marker.lower() in columns for marker in platform_markers):
                    return "platform"
                
                # 检测流水账单
                flow_markers = ["admin_id", "money", "delivery_order_id"]
                if any(marker.lower() in columns for marker in flow_markers):
                    return "flow"
                    
            except Exception:
                pass
        
        return None
    
    def parse_excel(
        self,
        file_id: uuid.UUID,
        header_row: int = 0,
    ) -> Tuple[Optional[List[Dict[str, Any]]], Optional[str]]:
        """
        解析 Excel 文件
        
        Returns:
            Tuple of (data list, error message)
        """
        file_record = self.get_file_by_id(file_id)
        if not file_record:
            return None, "文件不存在"
        
        file_path = file_record.file_path
        
        try:
            # 使用 pandas 读取
            df = pd.read_excel(
                file_path,
                header=header_row,
                dtype=str,  # 读取为字符串避免数字精度问题
            )
            
            # 转换为字典列表
            data = df.to_dict(orient="records")
            
            # 标记文件已处理
            file_record.is_processed = 1
            self.db.commit()
            
            return data, None
            
        except Exception as e:
            error_msg = f"解析失败: {str(e)}"
            file_record.parse_error = error_msg
            self.db.commit()
            return None, error_msg
    
    def parse_with_openpyxl(
        self,
        file_id: uuid.UUID,
        sheet_name: Optional[str] = None,
    ) -> Tuple[Optional[List[Dict[str, Any]]], Optional[str]]:
        """使用 openpyxl 解析 Excel (更精确的控制)"""
        file_record = self.get_file_by_id(file_id)
        if not file_record:
            return None, "文件不存在"
        
        file_path = file_record.file_path
        
        try:
            workbook = load_workbook(file_path, data_only=True)
            
            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                    return None, f"Sheet '{sheet_name}' 不存在"
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            # 获取表头
            headers = []
            for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)):
                headers.append(str(cell) if cell is not None else "")
            
            # 读取数据
            data = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                row_data = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_data[headers[i]] = value
                data.append(row_data)
            
            workbook.close()
            
            # 标记文件已处理
            file_record.is_processed = 1
            self.db.commit()
            
            return data, None
            
        except Exception as e:
            error_msg = f"解析失败: {str(e)}"
            file_record.parse_error = error_msg
            self.db.commit()
            return None, error_msg
    
    def list_files(
        self,
        user_id: str,
        file_type: Optional[str] = None,
        is_processed: Optional[bool] = None,
        skip: int = 0,
        limit: int = 20,
    ) -> Tuple[List[UploadedFile], int]:
        """获取文件列表"""
        query = self.db.query(UploadedFile).filter(UploadedFile.uploaded_by == uuid.UUID(user_id))
        
        if file_type:
            query = query.filter(UploadedFile.file_type == file_type)
        if is_processed is not None:
            query = query.filter(UploadedFile.is_processed == (1 if is_processed else 0))
            
        total = query.count()
        files = query.order_by(UploadedFile.created_at.desc()).offset(skip).limit(limit).all()
        
        return files, total

    def get_file_stats(self, file_id: uuid.UUID) -> Optional[Dict[str, Any]]:
        """获取文件统计信息"""
        file_record = self.get_file_by_id(file_id)
        if not file_record:
            return None
        
        return {
            "id": str(file_record.id),
            "filename": file_record.original_filename,
            "size": file_record.size_formatted,
            "type": file_record.file_type.value,
            "is_processed": bool(file_record.is_processed),
            "created_at": file_record.created_at.isoformat(),
        }
    
    def delete_file(self, file_id: uuid.UUID) -> bool:
        """删除文件"""
        file_record = self.get_file_by_id(file_id)
        if not file_record:
            return False
        
        # 删除物理文件
        try:
            if os.path.exists(file_record.file_path):
                os.remove(file_record.file_path)
        except Exception:
            pass
        
        # 删除数据库记录
        self.db.delete(file_record)
        self.db.commit()
        
        return True
    
    def _get_mime_type(self, filename: str) -> str:
        """获取 MIME 类型"""
        extension = Path(filename).suffix.lower()
        mime_types = {
            ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ".xls": "application/vnd.ms-excel",
            ".csv": "text/csv",
        }
        return mime_types.get(extension, "application/octet-stream")
